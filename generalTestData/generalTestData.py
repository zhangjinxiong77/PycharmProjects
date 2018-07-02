import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl import Workbook

# 自定义检测Excel的A列最大行函数。openpyxl的worksheet.max_row并不准
def find_maxrow(worksheet):
    i = 1
    while worksheet['A' + str(i)].value != None:
        i += 1
    return i-1

# testSet = "您好，请问您是机主本人吗M对啊M这里是玖富借贷服务平台，您在我平台申请的万卡借款已经逾期。想问一下，您是什么原因没有处理欠款呢M你听我讲M哦，不好意思，你请讲！"
dialogueTemplate = input("请输入对话模板：")
# replacedString = "你听我讲"
replacedString = input("请输入对话模板中需要替换的字符串：")
# wb = load_workbook('E:/4.corpus.xlsx')
excelPath = input("请输入存放语料的Excel的路径：")
wb = load_workbook(excelPath)
# ws = wb['Sheet1']
sheetName = input("请输入语料所在的Excel工作表名称：")
ws = wb[sheetName]
rows = find_maxrow(ws)
# 测试集需要存储到新的Excel中
wb_new = Workbook()
ws_new = wb_new.active


for row in range(2, rows):
    for col in range(1, 2):
        # print(ws.cell(column=col,row=row).value)
        newString: str = dialogueTemplate.replace(replacedString, ws.cell(column=col, row=row).value)
        # _ = ws.cell(column=col + 1, row=row, value="{0}".format(newString))
        ws_new.cell(column=col, row=row).value = newString
        print("当前处理到第"  + str(row) +  "条；共" + str(rows-1)+  "条" )

ws_new['A1'] = "模板"

fileName = input("请输入保存结果的Excel的名称和路径：")
# wb.save(filename='E:/4.testSet.xlsx')
wb_new.save(filename = fileName)


