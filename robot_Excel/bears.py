# -*- coding: utf-8 -*-

import requests
import json
import uuid
import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
import openpyxl.styles as sty

# 本脚本是库

# 寻找Excel的最后一行
def find_maxrow(worksheet):
    i = 1
    while worksheet['A' + str(i)].value != None:
        i += 1
    return i-1

# 寻找对话完整的CALL_ID,加入到一个List中。参数：ws（Excel工作表），rows(Excel工作表的最大行），callid（CALL_ID在第几列），dialogue（对话在第几列）
def integrated_callid(ws, rows, callid, dialogue):
    resultLi = []
    for row in range(2, rows + 1):
        for col in range(dialogue,dialogue+1):
            if "再见" in ws.cell(column=col, row=row).value:
                resultLi.append(ws.cell(column=callid, row=row).value)
    return resultLi

# 请求机器人的回复。参数：api_url（多轮对话接口的地址），para1（参数1），para2（参数2，指uuid）
def request_dialogue(api_url, para1, para2):
    dialogueRequest_url = api_url.format(para1, para2)
    dialogue_res = requests.get(dialogueRequest_url).text
    dialogue_res = json.loads(dialogue_res)
    return dialogue_res

# 请求意图。参数：url,uuid
def request_intention(api_url, uuid):
    intentionRequest_url = api_url.format(uuid)
    intention_res = requests.get(intentionRequest_url).text
    intention_res = json.loads(intention_res)
    return intention_res

