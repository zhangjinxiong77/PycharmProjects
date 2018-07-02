# -*- coding: utf-8 -*-

import time
import requests
import json
import re
import uuid
import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
import openpyxl.styles as sty

# 本脚本使用已生成的智能催收多轮对话测试集，调用多轮对话接口，来测试算法的准确性，同时给出对应的错误类型
# 自定义对话错误类型函数。参数为（正确回答和错误回答的合并，正确的回答，错误的回答）
def wrong_type(merge,rightAnswer,wrongAnswer):
    wrongType = ''
    # 流程的错误类型字典
    flow1 = "这里是玖富借贷服务平台，您在我平台申请的万卡借款已经逾期。想问一下，您是什么原因没有处理欠款呢and您好，是这样的，我们这边是玖富借贷服务平台，请您转告机主尽快联系万卡借贷服务平台，电话为’4008108818’，这个电话非常重要，麻烦您一定通知到他，我再重复一遍号码’4008108818’。或者请您报一下他的其他联系号码？"
    flow2 = "您好，是这样的，我们这边是玖富借贷服务平台，请您转告机主尽快联系万卡借贷服务平台，电话为’4008108818’，这个电话非常重要，麻烦您一定通知到他，我再重复一遍号码’4008108818’。或者请您报一下他的其他联系号码？and这里是玖富借贷服务平台，您在我平台申请的万卡借款已经逾期。想问一下，您是什么原因没有处理欠款呢"
    flow3 = "麻烦您履行承诺尽快处理您的欠款，若今晚24点之前您还没有处理您的欠款，明日将有专门的负责人员联系您，谢谢，再见and那您什么时候可以还款呢"
    flow4 = "那您什么时候可以还款呢and麻烦您履行承诺尽快处理您的欠款，若今晚24点之前您还没有处理您的欠款，明日将有专门的负责人员联系您，谢谢，再见"
    flow5 = "请您慎重考虑，尽快还款，否则我司有权启动仲裁程序追缴欠款，仲裁机构可依据您签署的调解协议依法作出裁决，并向法院申请强制执行。谢谢，再见and还希望您珍惜您的信用，尽快处理欠款，否则我司将采用仲裁等法律手段予以追偿，被仲裁后将列入法院’老赖’黑名单，会严重影响您的生活出行。还请您严肃对待，立即处理欠款，再见。"
    flow6 = "还希望您珍惜您的信用，尽快处理欠款，否则我司将采用仲裁等法律手段予以追偿，被仲裁后将列入法院’老赖’黑名单，会严重影响您的生活出行。还请您严肃对待，立即处理欠款，再见。and请您慎重考虑，尽快还款，否则我司有权启动仲裁程序追缴欠款，仲裁机构可依据您签署的调解协议依法作出裁决，并向法院申请强制执行。谢谢，再见"
    switcher1 = {
        flow1: "身份识别错误",
        flow2: "身份识别错误",
        flow3: "肯定识别为否定",
        flow4: "否定识别为肯定",
        flow5: "肯定识别为否定",
        flow6: "否定识别为肯定",
    }
    # 没有识别出的全局语境字典
    unGlobal1 = "这个不重要，打电话给你是为了让您重视您的贷款逾期问题，请你今天务必还清！"
    unGlobal2 = "我姓王，您叫我小王就可以了"
    unGlobal3 = "哦，不好意思，你请讲！"
    unGlobal4 = "您具体的逾期情况可在“玖富万卡”App和微信公众号里进行查看，也可以拨打客服电话400-810-8818进行咨询。"
    unGlobal5 = "我们的微信公众号和APP名称都是“玖富万卡”，您可以通过APP和公众号还款或查询账单。"
    unGlobal6 = "您可在“玖富万卡”App和微信公众号里进行查看还款。如果您对还款方式有疑问，直接拨打客服热线400-810-8818咨询。"
    unGlobal7 = "我们是玖富借贷服务平台的，你向我平台申请的万卡贷款已经逾期，打电话是提醒你，请您务必在今天24点前还清欠款的"
    unGlobal8 = "这个你只要抓紧把万卡借贷服务平台的欠款还清就行了，如果有问题可以联系我们客服热线400-810-8818，我再重复一遍号码：400-810-8818"
    unGlobal9 = "逾期不光会产生逾期利息，还可能会影响到您的个人名誉及信用等，建议您还是今天24点前偿还欠款吧，如果还有疑问可以拨打客服热线:400-810-8818咨询，行吧？"
    unGlobal10 = "请你务必在今天24点前还清万卡借贷服务平台的欠款，如果有任何疑问，可以拨打400-810-8818咨询。"
    unGlobal11 = "玖富借贷服务平台的业务是合法合规的，如果对我们的业务情况不满意，可拨打客服热线400-810-8818或拨打投诉热线400-810-1560进行投诉。"
    unGlobal12 = "您已经还掉了是吗？稍后我们这边也会再查看一下，打扰您了，再见。"
    switcher2 = {
        unGlobal1: "未识别出全局语境：询问是否机器人",
        unGlobal2: "未识别出全局语境：询问姓什么",
        unGlobal3: "未识别出全局语境：要求聆听",
        unGlobal4: "未识别出全局语境：询问逾期情况",
        unGlobal5: "未识别出全局语境：询问微信公众号",
        unGlobal6: "未识别出全局语境：询问还款方式",
        unGlobal7: "未识别出全局语境：询问是谁或来电目的",
        unGlobal8: "未识别出全局语境：询问客服热线",
        unGlobal9: "未识别出全局语境：协商或询问逾期后果",
        unGlobal10: "未识别出全局语境：询问还款时间",
        unGlobal11: "未识别出全局语境：投诉",
        unGlobal12: "未识别出：已还",
    }
    # 错误识别的全局语境字典
    switcher3 = {
        unGlobal1: "错误识别为：询问是否机器人",
        unGlobal2: "错误识别为：询问姓什么",
        unGlobal3: "错误识别为：要求聆听",
        unGlobal4: "错误识别为：询问逾期情况",
        unGlobal5: "错误识别为：询问微信公众号",
        unGlobal6: "错误识别为：询问还款方式",
        unGlobal7: "错误识别为：询问是谁或来电目的",
        unGlobal8: "错误识别为：询问客服热线",
        unGlobal9: "错误识别为：协商或询问逾期后果",
        unGlobal10: "错误识别为：询问还款时间",
        unGlobal11: "错误识别为：投诉",
        unGlobal12: "错误识别为：已还",
    }
    if merge in switcher1.keys():
        wrongType = wrongType + switcher1.get(merge) + "；"
    if rightAnswer in switcher2.keys():
        wrongType = wrongType + switcher2.get(rightAnswer) + "；"
    if wrongAnswer in switcher3.keys():
        wrongType = wrongType + switcher3.get(wrongAnswer) + "；"
    if wrongType != "":
        return wrongType
    else:
        return "其它错误类型"


# 自定义检测Excel的A列最大行函数。openpyxl的worksheet.max_row并不准
def find_maxrow(worksheet):
    i = 1
    while worksheet['A' + str(i)].value != None:
        i += 1
    return i-1


# 打开Excel
# wb = load_workbook('D:/已还测试集.xlsx')
# ws = wb['Sheet1']
excelPath = input("请输入Excel测试集的路径：")
wb = load_workbook(excelPath)
sheetName = input("请输入测试集数据所在的Excel工作表名称：")
ws = wb[sheetName]
# 注意：A列的第一行不得为空，否则rows=0
rows = find_maxrow(ws)

# 开始执行循环
for row in range(2, rows+1):
    # 测试数据必须放在A列
    for col in range(1, 2):
        # 调用多轮对话接口，并进行对话的正确性判断
        dialogue = ws.cell(column=col,row=row).value
        dialogue_list = dialogue.split('M')
        robot_talk = []
        # 注意：person_talk第1个为空值，用来触发多轮对话，“您好，请问您是机主本人吗？”是对该空值的回复
        person_talk = ['', ]
        for j in range(0, len(dialogue_list), 2):
            robot_talk.append(dialogue_list[j])
        for j in range(1, len(dialogue_list), 2):
            person_talk.append(dialogue_list[j])

        uuid_stamp = uuid.uuid4().hex
        contrast_log = []
        i = 0
        while i < len(person_talk):
            ask = person_talk[i]
            dialogueAPI_url = 'http://47.94.52.113:5061/api/v0.1/ask?project_id=48334234b5544627aa1cc977eec6cec6&query={}&uid={}'.format(ask, uuid_stamp)
            dialogue_res = requests.get(dialogueAPI_url).text
            dialogue_res = json.loads(dialogue_res)
            # 通过正则去掉接口返回的机器人回复的标点符号。以免测试集中的数据和接口返回的数据因为标点不同而导致的不匹配.
            answer = dialogue_res['data']['answer']
            trimed_answer = re.findall('\w', dialogue_res['data']['answer'])
            trimed_answer = ''.join(trimed_answer)
            # 通过正则去掉测试集中机器人回复的标点符号
            trimed_robotTalk = re.findall('\w', robot_talk[i])
            trimed_robotTalk = ''.join(trimed_robotTalk)
            # 对比
            if trimed_robotTalk == trimed_answer:
                i += 1
                print(ask)
                print(answer)
                contrast_log.append("客户:" + ask)
                contrast_log.append("机器人:" + answer)
            else:
                print(ask)
                print("不匹配的回答：" + trimed_answer)
                print("正确的回答：" + trimed_robotTalk)
                contrast_log.append("客户:" + ask)
                contrast_log.append("机器人（错误回答）:" + answer)
                contrast_log.append("机器人（正确回答）:" + robot_talk[i])
                # C列打印对话是否匹配
                ws.cell(column=col + 2, row=row).value = "对话不匹配"
                ws.cell(column=col + 2, row=row).fill = sty.PatternFill(fill_type='solid',fgColor="FFFF00")
                # D列打印对话的错误类型
                mergeAnswer = robot_talk[i] + "and" + dialogue_res['data']['answer']
                ws.cell(column=col + 3, row=row).value = wrong_type(mergeAnswer, robot_talk[i], answer)
                break
        # 打印对话记录。先将contast_log开头的 "" 空请求删除
        del contrast_log[0]
        contrast_log='\n'.join(contrast_log)
        # B列打印线上对话
        ws.cell(column=col + 1, row=row).value = contrast_log
        # 打印当前进度
        print("共" + str(rows-1) + "条；当前进度：%.2f%%" % (row/(rows-1)*100))

        # 调用意图接口并进行意图正确性的判断
        intention_url = 'http://47.95.36.52:13927/getdata?callId={}'.format(uuid_stamp)
        intention_res = requests.get(intention_url).text
        intention_res = json.loads(intention_res)
        intention_res = intention_res['data']['intention']
        # F列打印获取的线上意图
        ws.cell(column=col+5,row=row).value = intention_res
        # 正确的意图必须放在E列
        intention_right = ws.cell(column=col+4, row=row).value
        # G列打印意图是否匹配
        if intention_res != intention_right:
            ws.cell(column=col + 6, row=row).value = "意图不匹配"
            ws.cell(column=col + 6, row=row).fill =sty.PatternFill(fill_type='solid',fgColor="ADFF2F")


# 制作Excel的Title
ws['A1'] = "模板"
ws['B1'] = "线上对话"
ws['C1'] = "对话是否匹配"
ws['D1'] = "对话的错误类型"
ws['E1'] = "正确的意图"
ws['F1'] = "线上意图"
ws['G1'] = "意图是否匹配"

# 循环执行结束
fileName = input("请输入保存结果的Excel的名称和路径：")
wb.save(filename = fileName)




