# -*- coding: utf-8 -*-

import requests
import json
import uuid
import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
import openpyxl.styles as sty

# 本脚本使用已生成的智能催收多轮对话测试集，调用多轮对话接口，来测试算法的准确性，同时给出对应的错误类型
# 自定义检测Excel的A列最大行函数。openpyxl的worksheet.max_row并不准
def find_maxrow(worksheet):
    i = 1
    while worksheet['A' + str(i)].value != None:
        i += 1
    return i-1

# 打开Excel
wb = load_workbook('D:/智能催收测试集0713.xlsx')
ws = wb['Sheet1']
rows = find_maxrow(ws)

# 判断哪些Call_ID是对话完整的——将带有 “再见”的行ID加入List
completeLi = []
for row in range(2, rows + 1):
    # 对话数据必须放在D列.A列放CALL_ID
    for col in range(4, 5):
        if "再见" in ws.cell(column=col, row=row).value:
            completeLi.append(ws.cell(column=1, row=row).value)

# 创建字典。以compliteLi中的CALL_ID为key，以该轮对话的“正确/错误”为值
resultDic={}
for callid in completeLi:
    resultDic[callid] = ""

print('完整的对话共'+str(len(completeLi))+'条')

# 开始执行循环
for row in range(2, rows+1):
    for col in range(1, 2):
        print(row)
        if ws.cell(column=col,row=row).value in completeLi:
            # 如果是第一句话，则创建新的UUID，并发送''给接口，以开启多轮对话
            dialogue = ws.cell(column=col+3,row=row).value
            if dialogue == '机器人：您好，请问您是机主本人吗':
                uuid_stamp = uuid.uuid4().hex
                dialogueAPI_url = 'http://47.94.52.113:5061/api/v0.1/ask?project_id=48334234b5544627aa1cc977eec6cec6&query={}&uid={}'.format('', uuid_stamp)
                intention_url = 'http://47.95.36.52:13927/getdata?callId={}'.format(uuid_stamp)
                dialogue_res = requests.get(dialogueAPI_url).text
                print(dialogue)
            else:
                if dialogue[:2] == '客户':
                    if resultDic[ws.cell(column=col,row=row).value] != "错误" and resultDic[ws.cell(column=col,row=row).value] != "接口报错":
                        try:
                            # 获取机器人回复，并与测试集中的正确回复进行比较
                            slicedDialogue = dialogue[3:]
                            dialogueAPI_url = 'http://47.94.52.113:5061/api/v0.1/ask?project_id=48334234b5544627aa1cc977eec6cec6&query={}&uid={}'.format(slicedDialogue, uuid_stamp)
                            dialogue_res = requests.get(dialogueAPI_url).text
                            dialogue_res = json.loads(dialogue_res)
                            robotAnswer = dialogue_res['data']['answer']
                            robotAnswer = "机器人：" + robotAnswer
                            if robotAnswer != ws.cell(column=col+3, row=row+1).value:
                                resultDic[ws.cell(column=col,row=row).value] = "错误"
                                print(dialogue)
                                print('正确回答：'+ ws.cell(column=col+3, row=row+1).value)
                                print('错误回答：'+ robotAnswer)
                                ws.cell(column=col+5, row=row+1).value = robotAnswer
                            else:
                                print(dialogue)
                            # 获取意图，并与测试集中的正确意图进行比较
                            intention_res = requests.get(intention_url).text
                            intention_res = json.loads(intention_res)
                            intention_res = intention_res['data']['intention']
                            print("意图:" + intention_res)
                            if intention_res != ws.cell(column=col + 4, row=row).value:
                                ws.cell(column=col + 6, row=row).value = intention_res

                        except:
                            resultDic[ws.cell(column=col, row=row).value] = "接口报错"
                            ws.cell(column=col + 7, row=row).value = "接口报错"
                else:
                    if resultDic[ws.cell(column=col, row=row).value] != "错误" and resultDic[ws.cell(column=col,row=row).value] != "接口报错":
                        print(dialogue)

wb.save(filename='D:/测试集测试结果.xlsx')



