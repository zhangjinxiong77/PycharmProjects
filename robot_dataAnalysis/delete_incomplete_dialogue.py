# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
import datetime
import openpyxl.styles as sty
from openpyxl import Workbook

# 本脚本将对话分行的数据进行拼接，并标明是否对话完整。同时表明最后一句话术，以及对应的话术标签和挂断标签。
# 前提：A列：对话ID；B列：拨打时间；C列，姓名；D列，对话语句
def merge_dialogue():
    # 自定义检测Excel的A列最大行函数。openpyxl的worksheet.max_row并不准
    def find_maxrow(worksheet):
        i = 1
        while worksheet['A' + str(i)].value != None:
            i += 1
        return i - 1
    # 定义话术标签字典和挂断标签字典
    hangupDic = {
        "机器人：您好，请问您是机主本人吗": ["A", "询问是否机主本人", ],
        "机器人：这里是玖富借贷服务平台，您在我平台申请的万卡借款已经逾期。想问一下，您是什么原因没有处理欠款呢": ["B", "询问逾期原因", ],
        "机器人：您好，是这样的，我们这边是玖富借贷服务平台，请您转告机主尽快联系万卡借贷服务平台，电话为’4008108818’，这个电话非常重要，麻烦您一定通知到他，我再重复一遍号码’4008108818’。或者请您报一下他的其他联系号码？": ["C", "非本人", ],
        "机器人：您的情况我们了解了，但继续逾期会产生更多逾期费用，且会影响您的信用记录。那您今晚24点前能还款吗": ["D", "催还1", ],
        "机器人：麻烦您履行承诺尽快处理您的欠款，若今晚24点之前您还没有处理您的欠款，明日将有专门的负责人员联系您，谢谢，再见": ["E", "机器挂断1", ],
        "机器人：那您什么时候可以还款呢": ["F", "催还2", ],
        "机器人：请您慎重考虑，尽快还款，否则我司有权启动仲裁程序追缴欠款，仲裁机构可依据您签署的调解协议依法作出裁决，并向法院申请强制执行。谢谢，再见": ["G", "机器挂断2", ],
        "机器人：还希望您珍惜您的信用，尽快处理欠款，否则我司将采用仲裁等法律手段予以追偿，被仲裁后将列入法院’老赖’黑名单，会严重影响您的生活出行。还请您严肃对待，立即处理欠款，再见。": ["H", "机器挂断3", ],
        "机器人：谢谢您的配合，请您务必转告他，再见！": ["I", "转告", ],
        "机器人：您已经还掉了是吗？稍后我们这边也会再查看一下，打扰您了，再见。": ["J", "已还款", ],
        "机器人：这个不重要，打电话给你是为了让您重视您的贷款逾期问题，请你今天务必还清！": ["K", "全局：是否机器人", ],
        "机器人：我姓王，您叫我小王就可以了": ["L", "全局：姓什么", ],
        "机器人：哦，不好意思，你请讲！": ["M", "全局：要求聆听", ],
        "机器人：您具体的逾期情况可在“玖富万卡”App和微信公众号里进行查看，也可以拨打客服电话400-810-8818进行咨询。": ["N", "全局：询问逾期情况", ],
        "机器人：我们的微信公众号和APP名称都是“玖富万卡”，您可以通过APP和公众号还款或查询账单。": ["O", "全局：询问微信公众号", ],
        "机器人：您可在“玖富万卡”App和微信公众号里进行查看还款。如果您对还款方式有疑问，直接拨打客服热线400-810-8818咨询。": ["P", "全局：询问还款方式", ],
        "机器人：我们是玖富借贷服务平台的，你向我平台申请的万卡贷款已经逾期，打电话是提醒你，请您务必在今天24点前还清欠款的": ["Q", "全局：询问来电目的", ],
        "机器人：这个你只要抓紧把万卡借贷服务平台的欠款还清就行了，如果有问题可以联系我们客服热线400-810-8818，我再重复一遍号码：400-810-8818": ["R", "全局：询问联系方式", ],
        "机器人：逾期不光会产生逾期利息，还可能会影响到您的个人名誉及信用等，建议您还是今天24点前偿还欠款吧，如果还有疑问可以拨打客服热线:400-810-8818咨询，行吧？": ["S", "全局：客户协商/逾期后果", ],
        "机器人：请你务必在今天24点前还清万卡借贷服务平台的欠款，如果有任何疑问，可以拨打400-810-8818咨询。": ["T", "全局：时间关键词", ],
        "机器人：玖富借贷服务平台的业务是合法合规的，如果对我们的业务情况不满意，可拨打客服热线400-810-8818或拨打投诉热线400-810-1560进行投诉。": ["U", "全局：客户投诉", ],
    }

    # 打开Excel
    wb = load_workbook('D:/智能催收数据标注0627-0701.xlsx')
    ws = wb['Sheet1']
    rows = find_maxrow(ws)

    # 将带有 “再见”的行ID加入List
    completeLi= []
    for row in range(2, rows+1):
        # 对话数据必须放在D列.A列放
        for col in range(4, 5):
            if "再见" in ws.cell(column=col,row=row).value:
                completeLi.append(ws.cell(column=1,row=row).value)

    # 遍历Excel，如果对话ID在completeLi中，则将对话拼接入以对话ID为Key，List[对话时间，姓名，对话语句拼接，对话是否完整，挂断标签]为值的字典中。
    labelDataDic = {}
    for row in range(2, rows+1):
        for col in range(1, 2):
            if ws.cell(column=col,row=row).value in completeLi:
                # 如果对话ID已存在lableDataDic的Key中，则将“对话语句”前加换行符，接入该对话ID的记录中
                if ws.cell(column=col,row=row).value in labelDataDic.keys():
                    labelDataDic[ws.cell(column=col,row=row).value][2] += '\n' + ws.cell(column=col+3 ,row=row).value
                # 如果对话ID不在lableDataDic的Key中，则给该字典新增一条记录，以对话ID为Key，以List[对话时间，姓名，对话语句拼接，对话是否完整]为值。
                else:
                    labelDataDic[ws.cell(column=col,row=row).value] = [ws.cell(column=col+1, row=row).value.strftime('%Y-%m-%d %H:%M:%S'), ws.cell(column=col+2 ,row=row).value,ws.cell(column=col+3 ,row=row).value,"对话完整"]
            else:
                if ws.cell(column=col,row=row).value in labelDataDic.keys():
                    labelDataDic[ws.cell(column=col,row=row).value][2] += '\n' + ws.cell(column=col+3 ,row=row).value
                else:
                    labelDataDic[ws.cell(column=col,row=row).value] = [ws.cell(column=col+1, row=row).value.strftime('%Y-%m-%d %H:%M:%S'), ws.cell(column=col+2 ,row=row).value,ws.cell(column=col+3 ,row=row).value,"对话中断"]


    # print(labelDataDic)

    # 将字典存入Excel。因为字典无序不能循环，所以先将字典转换为List中的两位子tuple
    sumLi = []
    for kv in labelDataDic.items():
        sumLi.append(kv)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2['A1'] = "对话ID"
    ws2['B1'] = "拨打时间"
    ws2['C1'] = "客户姓名"
    ws2['D1'] = "对话语句"
    ws2['E1'] = "对话是否完整"
    ws2['F1'] = "被挂断的话术"
    ws2['G1'] = "话术标签"
    ws2['H1'] = "挂断标签"

    for row in range(2, len(labelDataDic) + 2):
        for col in range(1, 2):
            ws2.cell(column=col, row=row).value = sumLi[row - 2][0]
            ws2.cell(column=col + 1, row=row).value = sumLi[row - 2][1][0]
            ws2.cell(column=col + 2, row=row).value = sumLi[row - 2][1][1]
            ws2.cell(column=col + 3, row=row).value = sumLi[row - 2][1][2]
            ws2.cell(column=col + 4, row=row).value = sumLi[row - 2][1][3]
            # 添加挂断标签。无论对话是否完整
            dialogue_list = sumLi[row - 2][1][2].split('\n')
            ws2.cell(column=col + 5, row=row).value = dialogue_list[-1]
            print("第" + str(row-1) + "条：" + dialogue_list[-1])
            ws2.cell(column=col + 6, row=row).value = hangupDic.get(dialogue_list[-1])[0] if dialogue_list[-1] in hangupDic else "00000"
            ws2.cell(column=col + 7, row=row).value = hangupDic.get(dialogue_list[-1])[1] if dialogue_list[-1] in hangupDic else "00000"


    wb2.save(filename='D:/挂断标签统计.xlsx')


if __name__=='__main__':
    merge_dialogue()