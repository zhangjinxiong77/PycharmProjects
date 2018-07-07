# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

# 本脚本用来检测每个“类型”中，分数最高的用户都是哪个
def max_score():
    # 自定义检测Excel的A列最大行函数。openpyxl的worksheet.max_row并不准
    def find_maxrow(worksheet):
        i = 1
        while worksheet['A' + str(i)].value != None:
            i += 1
        return i - 1

    # 打开Excel。wb变量代表Excel工作薄，ws变量代表Excel工作表
    wb = load_workbook('D:/Python跑数.xlsx')
    ws = wb['Sheet5']
    rows = find_maxrow(ws)

    # 先定义一个空字典
    maxScoreDic = {}
    # 从第2行开始，对Excle中的每一行数据执行循环。前提：A列放类型，B列放用户，C列放分数
    for row in range(2, rows+1):
        for col in range(1, 2):
            # 如果该行的“类型”已经在字典的Key中，则判断该行的“分数”是否大于字典中该Key的分数，如果大于，则覆盖；
            if ws.cell(column=col,row=row).value in maxScoreDic.keys():
                if ws.cell(column=col+2,row=row).value > maxScoreDic[ws.cell(column=col, row=row).value][1]:
                    maxScoreDic[ws.cell(column=col, row=row).value][1] = float(ws.cell(column=col+2, row=row).value)
                    maxScoreDic[ws.cell(column=col, row=row).value][0] = ws.cell(column=col + 1, row=row).value
            # 如果该行的“类型”不在字典的Key中，则将改行的数据新增到字典中。以“类型”为Key，以列表[用户，分数]为Value
            else:
                maxScoreDic[ws.cell(column=col, row=row).value] = [ws.cell(column=col + 1, row=row).value, float(ws.cell(column=col + 2, row=row).value), ]

    # 将结果打印到控制台
    print(maxScoreDic)

    # 将结果打印到新的Excel工作薄。因为字典无序不能循环，所以先将字典转换为列表（List）
    sumLi = []
    for kv in maxScoreDic.items():
        sumLi.append(kv)

    wb2 = Workbook()
    ws2 = wb2.active
    # 打印Excel的Title
    ws2['A1'] = "类型"
    ws2['B1'] = "客户姓名"
    ws2['C1'] = "分数"
    for row in range(2, len(maxScoreDic) + 2):
        for col in range(1, 2):
            ws2.cell(column=col, row=row).value = sumLi[row - 2][0]
            ws2.cell(column=col+1, row=row).value = sumLi[row - 2][1][0]
            ws2.cell(column=col+2, row=row).value = sumLi[row - 2][1][1]

    wb2.save(filename='D:/最高分数统计Sheet5.xlsx')

if __name__=='__main__':
    max_score()