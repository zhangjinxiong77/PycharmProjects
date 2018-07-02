# -*- coding: utf-8 -*-

import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
import openpyxl.styles as sty
import random
from openpyxl import Workbook

# 定义话术列表
corpusLi = [
    "机器人：您好，请问您是机主本人吗",
    "机器人：这里是玖富借贷服务平台，您在我平台申请的万卡借款已经逾期。想问一下，您是什么原因没有处理欠款呢",
    "机器人：您好，是这样的，我们这边是玖富借贷服务平台，请您转告机主尽快联系万卡借贷服务平台，电话为’4008108818’，这个电话非常重要，麻烦您一定通知到他，我再重复一遍号码’4008108818’。或者请您报一下他的其他联系号码？",
    "机器人：您的情况我们了解了，但继续逾期会产生更多逾期费用，且会影响您的信用记录。那您今晚24点前能还款吗",
    "机器人：麻烦您履行承诺尽快处理您的欠款，若今晚24点之前您还没有处理您的欠款，明日将有专门的负责人员联系您，谢谢，再见",
    "机器人：那您什么时候可以还款呢",
    "机器人：请您慎重考虑，尽快还款，否则我司有权启动仲裁程序追缴欠款，仲裁机构可依据您签署的调解协议依法作出裁决，并向法院申请强制执行。谢谢，再见",
    "机器人：还希望您珍惜您的信用，尽快处理欠款，否则我司将采用仲裁等法律手段予以追偿，被仲裁后将列入法院’老赖’黑名单，会严重影响您的生活出行。还请您严肃对待，立即处理欠款，再见。",
    "机器人：谢谢您的配合，请您务必转告他，再见！",
    "机器人：您已经还掉了是吗？稍后我们这边也会再查看一下，打扰您了，再见。",
    "机器人：这个不重要，打电话给你是为了让您重视您的贷款逾期问题，请你今天务必还清！",
    "机器人：我姓王，您叫我小王就可以了",
    "机器人：哦，不好意思，你请讲！",
    "机器人：您具体的逾期情况可在“玖富万卡”App和微信公众号里进行查看，也可以拨打客服电话400-810-8818进行咨询。",
    "机器人：我们的微信公众号和APP名称都是“玖富万卡”，您可以通过APP和公众号还款或查询账单。",
    "机器人：您可在“玖富万卡”App和微信公众号里进行查看还款。如果您对还款方式有疑问，直接拨打客服热线400-810-8818咨询。",
    "机器人：我们是玖富借贷服务平台的，你向我平台申请的万卡贷款已经逾期，打电话是提醒你，请您务必在今天24点前还清欠款的",
    "机器人：这个你只要抓紧把万卡借贷服务平台的欠款还清就行了，如果有问题可以联系我们客服热线400-810-8818，我再重复一遍号码：400-810-8818",
    "机器人：逾期不光会产生逾期利息，还可能会影响到您的个人名誉及信用等，建议您还是今天24点前偿还欠款吧，如果还有疑问可以拨打客服热线:400-810-8818咨询，行吧？",
    "机器人：请你务必在今天24点前还清万卡借贷服务平台的欠款，如果有任何疑问，可以拨打400-810-8818咨询。",
    "机器人：玖富借贷服务平台的业务是合法合规的，如果对我们的业务情况不满意，可拨打客服热线400-810-8818或拨打投诉热线400-810-1560进行投诉。",
]

# 定义话术标签字典
corpusDic = {
    "机器人：您好，请问您是机主本人吗": 0,
    "机器人：这里是玖富借贷服务平台，您在我平台申请的万卡借款已经逾期。想问一下，您是什么原因没有处理欠款呢": 0,
    "机器人：您好，是这样的，我们这边是玖富借贷服务平台，请您转告机主尽快联系万卡借贷服务平台，电话为’4008108818’，这个电话非常重要，麻烦您一定通知到他，我再重复一遍号码’4008108818’。或者请您报一下他的其他联系号码？": 0,
    "机器人：您的情况我们了解了，但继续逾期会产生更多逾期费用，且会影响您的信用记录。那您今晚24点前能还款吗": 0,
    "机器人：麻烦您履行承诺尽快处理您的欠款，若今晚24点之前您还没有处理您的欠款，明日将有专门的负责人员联系您，谢谢，再见": 0,
    "机器人：那您什么时候可以还款呢": 0,
    "机器人：请您慎重考虑，尽快还款，否则我司有权启动仲裁程序追缴欠款，仲裁机构可依据您签署的调解协议依法作出裁决，并向法院申请强制执行。谢谢，再见": 0,
    "机器人：还希望您珍惜您的信用，尽快处理欠款，否则我司将采用仲裁等法律手段予以追偿，被仲裁后将列入法院’老赖’黑名单，会严重影响您的生活出行。还请您严肃对待，立即处理欠款，再见。": 0,
    "机器人：谢谢您的配合，请您务必转告他，再见！": 0,
    "机器人：您已经还掉了是吗？稍后我们这边也会再查看一下，打扰您了，再见。": 0,
    "机器人：这个不重要，打电话给你是为了让您重视您的贷款逾期问题，请你今天务必还清！": 0,
    "机器人：我姓王，您叫我小王就可以了": 0,
    "机器人：哦，不好意思，你请讲！": 0,
    "机器人：您具体的逾期情况可在“玖富万卡”App和微信公众号里进行查看，也可以拨打客服电话400-810-8818进行咨询。": 0,
    "机器人：我们的微信公众号和APP名称都是“玖富万卡”，您可以通过APP和公众号还款或查询账单。": 0,
    "机器人：您可在“玖富万卡”App和微信公众号里进行查看还款。如果您对还款方式有疑问，直接拨打客服热线400-810-8818咨询。": 0,
    "机器人：我们是玖富借贷服务平台的，你向我平台申请的万卡贷款已经逾期，打电话是提醒你，请您务必在今天24点前还清欠款的": 0,
    "机器人：这个你只要抓紧把万卡借贷服务平台的欠款还清就行了，如果有问题可以联系我们客服热线400-810-8818，我再重复一遍号码：400-810-8818": 0,
    "机器人：逾期不光会产生逾期利息，还可能会影响到您的个人名誉及信用等，建议您还是今天24点前偿还欠款吧，如果还有疑问可以拨打客服热线:400-810-8818咨询，行吧？": 0,
    "机器人：请你务必在今天24点前还清万卡借贷服务平台的欠款，如果有任何疑问，可以拨打400-810-8818咨询。": 0,
    "机器人：玖富借贷服务平台的业务是合法合规的，如果对我们的业务情况不满意，可拨打客服热线400-810-8818或拨打投诉热线400-810-1560进行投诉。": 0,
}


# 自定义检测Excel的A列最大行函数。openpyxl的worksheet.max_row并不准
def find_maxrow(worksheet):
    i = 1
    while worksheet['A' + str(i)].value != None:
        i += 1
    return i-1

# 打开Excel
wb = load_workbook('D:/对话语句导出0622-0624.xlsx')
ws = wb['Sheet1']
rows = find_maxrow(ws)

# 开始循环，计算每个话术出现的次数，即字典corpusDic的值
for row in range(2, rows+1):
    # 对话数据必须放在E列，并且以\n进行分隔
    for col in range(5, 6):
        dialogue = ws.cell(column=col, row=row).value
        dialogue_list = dialogue.split('\\n')
        for i in range(len(corpusLi)):
            if corpusLi[i] in dialogue_list:
                corpusDic[corpusLi[i]] += 1

# 生成新的工作薄和工作表，并将字典的值写入Excel中。因为字典没有顺序，所以先把字典转换为list中的两位子list或者子tuple。也可以转换成Pandas的DataFrame。
wb2 = Workbook()
ws2 = wb2.active

sumLi = []
for kv in corpusDic.items():
    sumLi.append(kv)

for row in range(1,len(corpusDic)+1):
    for col in range(1, 2):
        ws2.cell(column=col, row=row).value = sumLi[row - 1][0]
        ws2.cell(column=col+1, row=row).value = sumLi[row - 1][1]

wb2.save(filename='D:/话术频率统计0622-0624.xlsx')
