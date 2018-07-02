# -*- coding: utf-8 -*-
# 本脚本用来将数据库中导出的语料拆分到不同的Excel

import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl import Workbook

# 定义映射字典。k为在数据库中的标签；v为要保存的Excel的名称
mapDic = {
    '11': '语料库：已经还了',
    '21': '语料库：不是本人',
    '31': '语料库：承诺还款',
    '32': '语料库：积极协商',
    '41': '语料库：态度恶劣',
    '51': '语料库：其它',
    '61': '语料库：是本人',
    '847c465b051940f4a2cbac56f1610241': '语料库：是否机器人',
    '6ed79ec056c1418da908770ac094d90a': '语料库：姓什么',
    'e75d410f6b284aa296abdd88a6f28380': '要求聆听',
    'b9c7ea761e3b42c99e5c3e0f920e0ced': '询问逾期情况',
    '592f9d9485ec4463874a9ad69e48a2da': '询问微信公众号',
    'f7b4b265d56c4de884b3e9bd21eb068e': '询问还款方式',
    'c009ae0eb8984f238c16b13bb2fead0f': '询问来电目的',
    '998ab2998efd4290b675ed6254ec005c': '询问联系方式',
    '915b6787fa37448fa405d7a6a3d02f87': '客户协商/逾期后果',
    '8ed1bbfd91684bb59dcdf525c446203e': '时间关键词',
    '94bf960803aa4800ae7c2640012d7000': '客户投诉',
}

def txt2excel():
    f = open('D:/outbound.txt', 'r')
    lines = f.readlines()
    wb = Workbook()
    ws = wb.active

    for line in lines:
        line_list = line.split(',')
        row = lines.index(line)
        ws.cell(column=1, row=row+1).value = line_list[1]
        ws.cell(column=2, row=row+1).value = line_list[3]

    wb.save(filename='D:/数据库中的语料.xlsx')


if __name__=='__main__':
    txt2excel()
