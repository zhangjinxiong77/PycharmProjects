# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
import openpyxl.styles as sty

def find_maxrow ():
    wb = load_workbook('E:/testSet.xlsx')
    ws = wb['robot']
    i = 1
    while ws['A' + str(i)].value != None:
        print("当前进展到：" + str(i))
        i += 1
    return i-1


stra = find_maxrow()
print("最大行数为：" + str(stra))


